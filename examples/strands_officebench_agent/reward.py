"""OfficeBench reward function for evaluating task completion.

Adapts evaluation logic from OfficeBench's utils/evaluate.py to run locally
on the testbed directory. All evaluation is outcome-based: check files on disk.
"""

import ast
import difflib
import logging
import operator
import os
import re
from glob import glob

import icalendar
import openpyxl
import pytz

from agentcore_rl_toolkit import RewardFunction

logger = logging.getLogger(__name__)


# ── Safe comparator evaluation ──────────────────────────────────────────────
# OfficeBench task configs express a comparator as a Python lambda *string*
# (e.g. "lambda x: x in ['1', '2', '3']"). These configs are loaded from an S3
# URI supplied in the invocation payload, so they are untrusted input. Passing
# them to eval() is arbitrary code execution (RCE). Instead we parse the lambda
# with ast.parse() and evaluate it with a small allowlist interpreter: only a
# single-arg lambda whose body is built from comparisons (including membership),
# boolean/arithmetic/unary ops, numeric/string literals, list/tuple/set literals,
# the lambda's own parameter, and a handful of pure numeric-coercion builtins is
# permitted. Every other node type (attribute access, arbitrary names/calls,
# subscripts, comprehensions, imports, ...) is rejected before any code runs, so
# there is no reachable path to import, reflection, or the filesystem. ** is
# deliberately excluded to avoid cheap CPU/memory blow-ups (e.g. 2 ** 10**9).

# Builtins the comparator body may call. All are pure, side-effect-free coercions.
_SAFE_COMPARATOR_BUILTINS = {
    "int": int,
    "float": float,
    "str": str,
    "len": len,
    "abs": abs,
    "round": round,
    "bool": bool,
}

_SAFE_BINOPS = {
    ast.Add: operator.add,
    ast.Sub: operator.sub,
    ast.Mult: operator.mul,
    ast.Div: operator.truediv,
    ast.FloorDiv: operator.floordiv,
    ast.Mod: operator.mod,
}

_SAFE_COMPARES = {
    ast.Eq: operator.eq,
    ast.NotEq: operator.ne,
    ast.Lt: operator.lt,
    ast.LtE: operator.le,
    ast.Gt: operator.gt,
    ast.GtE: operator.ge,
    ast.In: lambda a, b: a in b,
    ast.NotIn: lambda a, b: a not in b,
}

_SAFE_UNARYOPS = {
    ast.UAdd: operator.pos,
    ast.USub: operator.neg,
    ast.Not: operator.not_,
}


class UnsafeComparatorError(ValueError):
    """Raised when a comparator string contains constructs outside the allowlist."""


def _make_safe_comparator(expr: str):
    """Compile an OfficeBench comparator lambda string into a safe callable.

    Accepts only a single-argument ``lambda`` whose body is composed of
    comparisons (including ``in``/``not in``), boolean/arithmetic/unary ops,
    numeric/string literals, list/tuple/set literals, the lambda parameter, and
    the coercion builtins in ``_SAFE_COMPARATOR_BUILTINS``. Raises
    ``UnsafeComparatorError`` for anything else, so a malicious task config can
    never execute arbitrary code.
    """
    try:
        tree = ast.parse(expr, mode="eval")
    except SyntaxError as e:
        raise UnsafeComparatorError(f"Comparator is not a valid expression: {expr!r}") from e

    if not isinstance(tree.body, ast.Lambda):
        raise UnsafeComparatorError(f"Comparator must be a lambda expression: {expr!r}")

    lam = tree.body
    args = lam.args
    if args.posonlyargs or args.kwonlyargs or args.vararg or args.kwarg or args.defaults or len(args.args) != 1:
        raise UnsafeComparatorError(f"Comparator lambda must take exactly one positional arg: {expr!r}")

    param_name = args.args[0].arg
    # Filled in per-call by the returned closure before evaluating the body.
    param_holder = {}

    def _eval(node):
        if isinstance(node, ast.Constant):
            if isinstance(node.value, (int, float, bool, str)):
                return node.value
            raise UnsafeComparatorError(f"Disallowed literal type: {type(node.value).__name__}")
        if isinstance(node, ast.Name):
            if node.id == param_name:
                return param_holder["value"]
            raise UnsafeComparatorError(f"Disallowed name reference: {node.id!r}")
        if isinstance(node, (ast.List, ast.Tuple, ast.Set)):
            elts = [_eval(e) for e in node.elts]
            if isinstance(node, ast.List):
                return elts
            if isinstance(node, ast.Tuple):
                return tuple(elts)
            return set(elts)
        if isinstance(node, ast.BinOp) and type(node.op) in _SAFE_BINOPS:
            return _SAFE_BINOPS[type(node.op)](_eval(node.left), _eval(node.right))
        if isinstance(node, ast.UnaryOp) and type(node.op) in _SAFE_UNARYOPS:
            return _SAFE_UNARYOPS[type(node.op)](_eval(node.operand))
        if isinstance(node, ast.BoolOp):
            values = [_eval(v) for v in node.values]
            if isinstance(node.op, ast.And):
                return all(values)
            return any(values)
        if isinstance(node, ast.Compare):
            left = _eval(node.left)
            result = True
            for op, comparator_node in zip(node.ops, node.comparators, strict=True):
                op_type = type(op)
                if op_type not in _SAFE_COMPARES:
                    raise UnsafeComparatorError(f"Disallowed comparison operator: {op_type.__name__}")
                right = _eval(comparator_node)
                if not _SAFE_COMPARES[op_type](left, right):
                    result = False
                    break
                left = right
            return result
        if isinstance(node, ast.Call):
            if not isinstance(node.func, ast.Name) or node.func.id not in _SAFE_COMPARATOR_BUILTINS or node.keywords:
                raise UnsafeComparatorError("Only calls to allowlisted numeric builtins are permitted")
            fn = _SAFE_COMPARATOR_BUILTINS[node.func.id]
            return fn(*[_eval(a) for a in node.args])
        raise UnsafeComparatorError(f"Disallowed expression: {type(node).__name__}")

    def comparator(value):
        param_holder["value"] = value
        return bool(_eval(lam.body))

    return comparator


# ── File reading helpers (adapted from OfficeBench apps) ────────────────────


def _read_excel(file_path: str) -> str:
    """Read Excel file into '(row, col): value' format."""
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active
    result = ""
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value is not None:
                result += f"({cell.row}, {cell.column}): {cell.value}\t"
        result += "\n"
    return result


def _read_word(file_path: str) -> str:
    """Read Word file into string."""
    from docx import Document

    doc = Document(file_path)
    return "\n".join(p.text for p in doc.paragraphs)


def _read_pdf(file_path: str) -> str:
    """Read PDF file into string."""
    import fitz

    doc = fitz.open(file_path)
    text = ""
    for page in doc:
        text += page.get_text()
    doc.close()
    return text


def _read_pdf_pages(file_path: str) -> list:
    """Read PDF file into list of page texts."""
    import fitz

    doc = fitz.open(file_path)
    pages = []
    for page in doc:
        pages.append(page.get_text())
    doc.close()
    return pages


def _list_emails(username: str, testbed_dir: str, first_n_letters: int = -1) -> str:
    """List emails for a given username."""
    from email import policy
    from email.parser import BytesParser

    email_folder = os.path.join(testbed_dir, "emails", username)
    if not os.path.exists(email_folder):
        return ""
    email_files = glob(os.path.join(email_folder, "*.eml"))
    message = ""
    for email_file in email_files:
        with open(email_file, "rb") as f:
            email_content = f.read()
            email = BytesParser(policy=policy.default).parsebytes(email_content)
        email_name = os.path.basename(email_file)
        message += f"Email ID: {email_name}\n"
        message += f"From: {email['From']}\n"
        message += f"To: {email['To']}\n"
        message += f"Subject: {email['Subject']}\n"

        # Get content
        if email.is_multipart():
            parts = []
            for part in email.iter_parts():
                ct = part.get_content_type()
                if ct in ("text/plain", "text/html"):
                    parts.append(
                        part.get_payload(decode=True).decode(part.get_content_charset() or "utf-8", errors="replace")
                    )
            content = "\n".join(parts)
        else:
            content = email.get_payload(decode=True).decode(email.get_content_charset() or "utf-8", errors="replace")

        if first_n_letters != -1:
            message += f"Content: {content[:first_n_letters]}...\n"
        else:
            message += f"Content: {content}\n"
        message += "-" * 50 + "\n"
    return message


# ── Evaluation functions ────────────────────────────────────────────────────


def _is_number(string: str) -> bool:
    try:
        float(string)
        return True
    except ValueError:
        return False


def _evaluate_contain_text(content: str, args: dict) -> bool:
    content = content.lower()
    for keyword in args["keywords"]:
        keyword = keyword.lower()
        if _is_number(keyword):
            content = content.replace(",", "")
        if keyword not in content:
            return False
    return True


def evaluate_contain(testbed_dir: str, args: dict) -> bool:
    doc_type = args["doc_type"]
    if doc_type == "email":
        username = args["username"]
        email_contents = ""
        email_dir = os.path.join(testbed_dir, "emails")
        if os.path.exists(os.path.join(email_dir, username)):
            email_contents = _list_emails(username, testbed_dir, -1)
        elif os.path.exists(os.path.join(email_dir, username.lower())):
            email_contents = _list_emails(username.lower(), testbed_dir, -1)
        else:
            # Try case-insensitive match
            if os.path.exists(email_dir):
                for account in os.listdir(email_dir):
                    if username.lower() in account.lower():
                        email_contents = _list_emails(account, testbed_dir, -1)
                        break
        return _evaluate_contain_text(email_contents, args)

    file_path = os.path.join(testbed_dir, args["file"])
    if not os.path.exists(file_path) or not os.path.isfile(file_path):
        logger.warning(f"File not found: {file_path}")
        return False

    if doc_type == "xlsx":
        content = _read_excel(file_path)
    elif doc_type in ("txt", "ics"):
        with open(file_path) as f:
            content = f.read()
    elif doc_type in ("doc", "docx"):
        content = _read_word(file_path)
    elif doc_type == "pdf":
        content = _read_pdf(file_path)
    else:
        logger.error(f"Unsupported doc type: {doc_type}")
        return False

    return _evaluate_contain_text(content, args)


def evaluate_not_contain(testbed_dir: str, args: dict) -> bool:
    return not evaluate_contain(testbed_dir, args)


def evaluate_file_exist(testbed_dir: str, args: dict) -> bool:
    return os.path.exists(os.path.join(testbed_dir, args["file"]))


def evaluate_file_not_exist(testbed_dir: str, args: dict) -> bool:
    return not os.path.exists(os.path.join(testbed_dir, args["file"]))


def evaluate_diff_contain_text(testbed_dir: str, args: dict) -> bool:
    doc_type = args["doc_type"]
    input_file = os.path.join(testbed_dir, args["input_file"])
    output_file = os.path.join(testbed_dir, args["output_file"])

    if doc_type == "xlsx":
        helper = _read_excel
    elif doc_type == "doc":
        helper = _read_word
    else:
        logger.error(f"Unsupported doc type for diff: {doc_type}")
        return False

    input_content = helper(input_file)
    output_content = helper(output_file)

    if input_content == output_content:
        return False

    diff = "\n".join(difflib.unified_diff(input_content.split("\n"), output_content.split("\n"), n=0))
    for keyword in args["keywords"]:
        if keyword not in diff:
            return False
    return True


def evaluate_excel_cell_value(testbed_dir: str, args: dict) -> bool:
    file_path = os.path.join(testbed_dir, args["file"])
    if not os.path.exists(file_path):
        logger.warning(f"File not found: {file_path}")
        return False

    content = _read_excel(file_path)
    for match in args["matches"]:
        pattern = f"({match['row']}, {match['col']}): {match['value']}"
        if pattern not in content:
            return False
    return True


def evaluate_excel_cell_comparator(testbed_dir: str, args: dict) -> bool:
    file_path = os.path.join(testbed_dir, args["file"])
    if not os.path.exists(file_path):
        logger.warning(f"File not found: {file_path}")
        return False

    content = _read_excel(file_path)
    for match in args["matches"]:
        pattern = r"\({}, {}\): (\w+)\t".format(match["row"], match["col"])
        x = re.search(pattern, content)
        if x:
            value = x.group(1)
            try:
                comparator = _make_safe_comparator(match["comparator"])
            except UnsafeComparatorError as e:
                # Untrusted comparator outside the safe allowlist — never execute it.
                logger.warning("Rejected unsafe comparator %r: %s", match.get("comparator"), e)
                return False
            if comparator(value):
                continue
            else:
                return False
        else:
            return False
    return True


def evaluate_calendar_no_overlap(testbed_dir: str, args: dict) -> bool:
    username = args["username"]
    calendar_file = os.path.join(testbed_dir, "calendar", f"{username}.ics")
    if not os.path.exists(calendar_file):
        logger.warning(f"Calendar file not found: {calendar_file}")
        return False

    with open(calendar_file, "rb") as f:
        calendar = icalendar.Calendar.from_ical(f.read())

    utc = pytz.UTC

    def proc_dt(dt):
        if dt.tzinfo is None:
            return utc.localize(dt)
        return dt

    calendar.subcomponents.sort(key=lambda x: proc_dt(x.get("dtstart").dt))

    events = [c for c in calendar.walk() if c.name == "VEVENT"]
    for i in range(len(events) - 1):
        if proc_dt(events[i].get("dtend").dt) > proc_dt(events[i + 1].get("dtstart").dt):
            return False
    return True


def evaluate_exact_match(testbed_dir: str, args: dict) -> bool:
    result_path = os.path.join(testbed_dir, args["result_file"])
    if not os.path.exists(result_path):
        logger.warning(f"Result file not found: {result_path}")
        return False

    expected_path = os.path.join(testbed_dir, args["expected_file"])
    doc_type = args["doc_type"]

    if doc_type == "xlsx":
        result_sheet = openpyxl.load_workbook(result_path).active
        expected_sheet = openpyxl.load_workbook(expected_path).active

        for row in result_sheet.iter_rows():
            for cell in row:
                expected_value = expected_sheet.cell(row=cell.row, column=cell.column).value
                if cell.value != expected_value:
                    return False
        for row in expected_sheet.iter_rows():
            for cell in row:
                result_value = result_sheet.cell(row=cell.row, column=cell.column).value
                if cell.value != result_value:
                    return False
        return True

    def _read_text_file(x):
        return open(x).read()

    if doc_type in ("txt", "ics"):
        helper = _read_text_file
    elif doc_type == "doc":
        helper = _read_word
    elif doc_type == "pdf":
        helper = _read_pdf_pages
    else:
        logger.error(f"Unsupported doc type for exact match: {doc_type}")
        return False

    return helper(result_path) == helper(expected_path)


# ── Evaluation function registry ────────────────────────────────────────────

EVAL_FUNCTIONS = {
    "evaluate_contain": evaluate_contain,
    "evaluate_not_contain": evaluate_not_contain,
    "evaluate_file_exist": evaluate_file_exist,
    "evaluate_file_not_exist": evaluate_file_not_exist,
    "evaluate_diff_contain_text": evaluate_diff_contain_text,
    "evaluate_excel_cell_value": evaluate_excel_cell_value,
    "evaluate_excel_cell_comparator": evaluate_excel_cell_comparator,
    "evaluate_calendar_no_overlap": evaluate_calendar_no_overlap,
    "evaluate_exact_match": evaluate_exact_match,
}


# ── Reward Function ─────────────────────────────────────────────────────────


def _normalize_eval_path(path: str) -> str:
    """Normalize file paths in evaluation args.

    OfficeBench eval configs contain paths like:
      - ../../../../reference/score.xlsx  -> reference/score.xlsx
      - ../../../../cache/0/testbed/data/score.xlsx -> cache/data/score.xlsx

    These are relative to OfficeBench's Docker output structure. We normalize
    them to resolve within our flat /testbed/ directory.
    """
    if "../../../../reference/" in path:
        # ../../../../reference/foo.xlsx -> reference/foo.xlsx
        return "reference/" + path.split("../../../../reference/")[-1]
    if "../../../../cache/" in path:
        # ../../../../cache/{id}/testbed/data/foo.xlsx -> cache/data/foo.xlsx
        # Strip the cache/{id}/testbed/ prefix, keep the rest under cache/
        after_cache = path.split("../../../../cache/")[-1]
        # after_cache = "0/testbed/data/score.xlsx"
        parts = after_cache.split("/", 2)  # ["0", "testbed", "data/score.xlsx"]
        if len(parts) >= 3:
            return "cache/" + parts[2]
        return "cache/" + after_cache
    return path


def _normalize_eval_args(args: dict) -> dict:
    """Normalize all file path values in eval args."""
    path_keys = ["file", "input_file", "output_file", "result_file", "expected_file"]
    normalized = dict(args)
    for key in path_keys:
        if key in normalized and isinstance(normalized[key], str):
            normalized[key] = _normalize_eval_path(normalized[key])
    return normalized


class OfficeBenchReward(RewardFunction):
    def __call__(self, testbed_dir: str, evaluation_config: list, **kwargs) -> float:
        """Run all evaluation checks from the task config.

        Returns 1.0 if ALL checks pass, 0.0 otherwise.
        """
        for eval_item in evaluation_config:
            function_name = eval_item["function"]
            args = _normalize_eval_args(eval_item["args"])

            eval_fn = EVAL_FUNCTIONS.get(function_name)
            if eval_fn is None:
                logger.error(f"Unknown evaluation function: {function_name}")
                return 0.0

            try:
                if not eval_fn(testbed_dir, args):
                    logger.info(f"Evaluation check failed: {function_name} with args {args}")
                    return 0.0
            except Exception:
                logger.exception(f"Error running evaluation function: {function_name}")
                return 0.0

        logger.info("All evaluation checks passed!")
        return 1.0


if __name__ == "__main__":
    reward_fn = OfficeBenchReward()
    # Example usage:
    # reward = reward_fn(
    #     testbed_dir="/testbed",
    #     evaluation_config=[
    #         {"function": "evaluate_file_exist", "args": {"file": "data/output.xlsx"}}
    #     ],
    # )
    # print(reward)
